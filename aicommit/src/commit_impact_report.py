import os
import subprocess
import requests
import json
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from io import BytesIO
import gspread
from oauth2client.service_account import ServiceAccountCredentials
 
# Load environment variables from .env file
load_dotenv()
 
# Azure OpenAI Setup
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
AZURE_API_ENDPOINT = "https://shrutiaiinstance.openai.azure.com/openai/deployments/gpt-4o-mini/chat/completions?api-version=2025-01-01-preview"
HEADERS = {
    "Content-Type": "application/json",
    "api-key": OPENAI_API_KEY
}
 
# Microsoft Graph Auth Setup (future use)
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
EXCEL_FILE_PATH_ONEDRIVE = "/drive/root:/PR_Report.xlsx"
 
 
# Function to get an access token from Microsoft Graph
def get_access_token():
    token_url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    payload = {
        'client_id': CLIENT_ID,
        'scope': 'https://graph.microsoft.com/.default',
        'client_secret': CLIENT_SECRET,
        'grant_type': 'client_credentials'
    }
    response = requests.post(token_url, data=payload)
    response.raise_for_status()  # Ensure we raise exception for bad requests
    return response.json()["access_token"]
 
 
# Function to download the Excel file from OneDrive (for future use)
def download_excel_file(access_token):
    url = f"https://graph.microsoft.com/v1.0{EXCEL_FILE_PATH_ONEDRIVE}:/content"
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(url, headers=headers)
    # Return file content in BytesIO if successful, else None
    return BytesIO(response.content) if response.status_code == 200 else None
 
 
# Function to upload the Excel file back to OneDrive (for future use)
def upload_excel_file(access_token, file_content):
    url = f"https://graph.microsoft.com/v1.0{EXCEL_FILE_PATH_ONEDRIVE}:/content"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    response = requests.put(url, headers=headers, data=file_content)
    if response.status_code in [200, 201]:
        print("✅ Excel file updated on OneDrive.")
    else:
        print(f"❌ Failed to upload Excel file: {response.status_code}")
        print(response.text)
 
 
# Append data to a local Excel file stored in the .github folder
def append_to_excel_local(date, username, commit_message, module_summary):
    local_excel_path = ".github/PR_Report.xlsx"
    # Check if the file exists and load it, else create a new workbook
    if os.path.exists(local_excel_path):
        workbook = load_workbook(local_excel_path)
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Add header if new workbook created
        sheet.append(["Date", "GitHub Username", "Commit Message", "Impacted Backend Modules"])
 
    sheet = workbook.active
    # Append the new row with commit details
    sheet.append([date, username, commit_message, module_summary])
    workbook.save(local_excel_path)
 
 
# Retrieve git diff between master and current HEAD
def get_commit_diff():
    try:
        diff = subprocess.check_output(
            ["git", "diff", "master..HEAD", "--no-color"],
            stderr=subprocess.DEVNULL
        )
        # Return git diff as a string
        return diff.decode("utf-8").strip()
    except subprocess.CalledProcessError as e:
        print("❌ Error getting git diff:", e)
        return ""
 
 
# Generate a full impact report using Azure OpenAI API
def generate_full_impact_report(diff_text):
    prompt = f"""
    You're an expert code reviewer and software architect.
 
    Given the following git diff from a project that may include both backend and frontend code, perform an Impact Area Analysis Report. Include:
 
    - APIs Affected
    - Modules Changed (Backend)
    - Frontend Changes
    - Reasoning
 
    Git Diff:
    {diff_text}
    """
    payload = {
        "messages": [
            {"role": "system", "content": "You are a senior developer reviewing code changes."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.5,
        "max_tokens": 1000
    }
    response = requests.post(AZURE_API_ENDPOINT, headers=HEADERS, data=json.dumps(payload))
    if response.status_code == 200:
        # Return the response text from the AI
        return response.json()["choices"][0]["message"]["content"].strip()
    else:
        print("❌ Azure OpenAI API Error:", response.status_code, response.text)
        return ""
 
 
# Generate a summary by extracting only backend module names via Azure OpenAI API
def extract_backend_modules(diff_text):
    prompt = f"""
    You're a backend expert. From the git diff below, extract ONLY backend module names (e.g., file names, classes, packages, functions) that were affected. Return a clean comma-separated list.
 
    Git Diff:
    {diff_text}
    """
    payload = {
        "messages": [
            {"role": "system", "content": "You extract backend modules from code diffs."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.3,
        "max_tokens": 300
    }
    response = requests.post(AZURE_API_ENDPOINT, headers=HEADERS, data=json.dumps(payload))
    if response.status_code == 200:
        # Return the clean, comma-separated list of backend modules
        return response.json()["choices"][0]["message"]["content"].strip()
    else:
        print("❌ Azure OpenAI API Error (backend modules):", response.status_code, response.text)
        return ""
 
 
# Append a row of commit data to Google Sheets
def append_to_google_sheet(date, username, commit_message, module_summary):
    # Define the scope for Google Sheets API
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive"
    ]
 
    # Path to your JSON key file (downloaded from Google Cloud Console)
    CREDENTIALS_PATH = os.path.join(os.path.dirname(__file__), "google_credentials.json")
    creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_PATH, scope)
 
    # Authorize the client with the credentials
    client = gspread.authorize(creds)
 
    # Open the sheet by its key
    sheet = client.open_by_key("1D4pbXFL6Uh6Zu1W7ZPfHFuHzy96Fx5Ok7_pKFfEVwGY").sheet1
    # Append the row with commit data
    sheet.append_row([date, username, commit_message, module_summary])
 
    print("✅ Summary added to Google Sheet.")
 
 
# Retrieve the latest commit message from git
def get_commit_message():
    try:
        return subprocess.check_output(["git", "log", "-1", "--pretty=%B"]).decode("utf-8").strip()
    except subprocess.CalledProcessError:
        return "N/A"
 
 
# Retrieve the git username configured locally
def get_git_username():
    try:
        return subprocess.check_output(["git", "config", "user.name"]).decode("utf-8").strip()
    except subprocess.CalledProcessError:
        return "N/A"
 
 
# Main function to generate the impact report, save it and update summary
def impact_report():
    if not OPENAI_API_KEY:
        print("❌ OPENAI_API_KEY not set in .env")
        return
 
    print("📦 Fetching latest commit diff...")
    diff = get_commit_diff()
    if not diff:
        print("⚠️ No changes detected.")
        return
 
    print("🧠 Generating full Impact Area Analysis Report...")
    full_report = generate_full_impact_report(diff)
 
    print("📦 Extracting only backend module names for summary...")
    backend_modules = extract_backend_modules(diff)
 
    # Save full report to markdown file inside .github folder
    if full_report:
        os.makedirs(".github", exist_ok=True)
        with open(".github/IMPACT_REPORT.md", "w", encoding="utf-8") as f:
            f.write(full_report)
        print("💾 Full impact report saved to `.github/IMPACT_REPORT.md`")
 
    # Prepare commit metadata details
    date = datetime.now().strftime("%Y-%m-%d %H:%M")
    username = get_git_username()
    commit_message = get_commit_message()
    
    # Append summary to the storage
    # Uncomment below to use local Excel; currently using Google Sheets
    # append_to_excel_local(date, username, commit_message, backend_modules)
    append_to_google_sheet(date, username, commit_message, backend_modules)
    print("✅ Summary added to `.github/PR_Report.xlsx`")